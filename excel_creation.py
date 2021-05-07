import pandas as pd
from openpyxl import load_workbook, Workbook

pd.options.display.float_format = '{:.0f}'.format


class OutputCreator:
    def __init__(self, soups, ticker):
        self.soups = soups
        self.ticker = ticker
        self.df_list = []
        self.req_type_list = ['financials', 'balance-sheet', 'cash-flow']

    @staticmethod
    def find_content(soup):
        content = []

        raw_content = soup.find_all('div', class_='D(tbrg)')
        for item in raw_content[0].find_all('div', class_='D(tbc)'):
            item = item.text.replace(',', '')
            try:
                item = int(item)
            except ValueError:
                pass
            content.append(item)

        return content

    @staticmethod
    def divide_content_by_row(headers, content):
        nb_col = len(headers)
        row_content = []
        for i in range(0, int(len(content) / nb_col)):
            row_content.append(content[i * nb_col:(i + 1) * nb_col])

        return row_content

    def create_dataframe(self):

        for soup in self.soups:

            headers = ['Breakdown']
            titles = soup.find_all('div', class_='D(tbr)')
            for item in titles[0].find_all('div', class_='Ta(c)'):
                headers.append(item.text)

            content = self.find_content(soup)
            row_content = self.divide_content_by_row(headers, content)

            df = pd.DataFrame(row_content)
            df.columns = headers

            self.df_list.append(df)

    def df_to_excel(self):
        path = 'output.xlsx'
        try:
            book = load_workbook(path)
        except FileNotFoundError:
            book = Workbook()
            book.active.title = self.ticker
            book.save(filename=path)

        with pd.ExcelWriter(path) as writer:
            writer.book = book

            df_final = pd.DataFrame()
            for index, req_type in enumerate(self.req_type_list):
                # Add a column with the name of the financial statement from where the data is coming from
                self.df_list[index].insert(0, 'Statement', req_type)
                df_final = df_final.append(self.df_list[index], sort=False, ignore_index=True)

            try:  # Remove sheet if already in workbook
                std = book.get_sheet_by_name(self.ticker)
                book.remove_sheet(std)
            except KeyError:
                pass

            df_final.to_excel(writer, sheet_name=self.ticker)
